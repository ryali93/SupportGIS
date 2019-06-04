#!/usr/bin/env python
# -*- coding: utf-8 -*-

import arcpy, os
import math
import pandas as pd
from datetime import datetime
import sys
from openpyxl import load_workbook
import traceback

startTime = datetime.now()
print startTime

arcpy.env.overwriteOutput = True

BASE_DIR = os.path.abspath(os.path.join(__file__, '..'))

pathGDB = os.path.join(BASE_DIR, 'MODELO_CLIENTES.gdb')
TAP     = os.path.join(pathGDB, 'TAP')
NAP     = os.path.join(pathGDB, 'NAP')
CLIENTE = os.path.join(pathGDB, 'CLIENTE')
AINFL   = os.path.join(pathGDB, 'AREA_INFL')
ACONTR  = os.path.join(pathGDB, 'AREA_CONTR')

AINFL_TEMP  = os.path.join(pathGDB, 'AINFL_TEMP')
NAP_TEMP    = os.path.join(pathGDB, 'NAP_TEMP')
TAP_TEMP    = os.path.join(pathGDB, 'TAP_TEMP')

pathExcel   = os.path.join(BASE_DIR, 'CABECERAS.xlsx')
pathExcelout= os.path.join(BASE_DIR, 'CABECERAS_1.xlsx')

codTAP      = "COD_TAP"
codNAP      = "mn_pky_nap"

ACONT_mfl   = arcpy.MakeFeatureLayer_management(AINFL, "ACONT_mfl")
AINFL_mfl   = arcpy.MakeFeatureLayer_management(AINFL, "AINFL_mfl")
TAP_mfl     = arcpy.MakeFeatureLayer_management(TAP, "TAP_mfl")
NAP_mfl     = arcpy.MakeFeatureLayer_management(NAP, "NAP_mfl")


def selectLayer(lyr, ws):
    cantCli = int(arcpy.GetCount_management(lyr).getOutput(0))
    container = []
    reg = 2
    for i in range(cantCli)[:200000]:
        try:
            cli   = arcpy.MakeFeatureLayer_management(CLIENTE, "cli", "OBJECTID = %s" % i)
            acont = arcpy.SelectLayerByLocation_management("ACONT_mfl", 'INTERSECT', "cli", '#', 'NEW_SELECTION', 'NOT_INVERT')  # Seleccionar capa de Area de influencia de cada cliente
            ainfl = arcpy.SelectLayerByLocation_management("AINFL_mfl", 'INTERSECT', "cli", '#', 'NEW_SELECTION', 'NOT_INVERT')      # Seleccionar capa de Area de influencia de cada cliente
            napfl = arcpy.SelectLayerByLocation_management("NAP_mfl", 'INTERSECT', acont, '#', 'NEW_SELECTION', 'NOT_INVERT')        # Seleccionar capas de NAP
            tapfl = arcpy.SelectLayerByLocation_management("TAP_mfl", 'INTERSECT', ainfl, '#', 'NEW_SELECTION', 'NOT_INVERT')        # Seleccionar capas de TAP

            with arcpy.da.SearchCursor(cli, ["ID", "SHAPE@X", "SHAPE@Y"]) as cursor:
                for x in cursor:
                    ws["A{}".format(reg)] = x[0]
                    ws["B{}".format(reg)] = x[1]
                    ws["C{}".format(reg)] = x[2]
                    ws["M{}".format(reg)] = "NO"

            listaNAP = []
            with arcpy.da.SearchCursor(napfl, [codNAP, "SHAPE@X", "SHAPE@Y", "mn_estado_nap", "mn_capacidad_nap", "mn_cnt_hilos_libres","mn_tipo_nap", "mn_sector_tdp","mn_numcoo_x", "mn_numcoo_y"]) as cursor:
                for m in cursor:
                    cli_fd = [[x[0], x[1], x[2]] for x in arcpy.da.SearchCursor(cli, ["ID", "SHAPE@X", "SHAPE@Y"])]
                    cli_x  = cli_fd[0][1]
                    cli_y  = cli_fd[0][2]
                    nap_x  = m[1]
                    nap_y  = m[2]
                    dist = round(math.sqrt((pow(cli_x - nap_x, 2) + pow(cli_y - nap_y,2)))*111110, 1)
                    rowNAP = [dist, m[0], m[3], m[4], m[5], m[6], m[7], m[8], m[9], "SI"]
                    listaNAP.append(rowNAP)
            listaNAP.sort(key=lambda x:x[0])
            if len(listaNAP) != 0:
                ws["D{}".format(reg)] = listaNAP[0][0]
                ws["E{}".format(reg)] = listaNAP[0][1]
                ws["F{}".format(reg)] = listaNAP[0][2]
                ws["G{}".format(reg)] = listaNAP[0][3]
                ws["H{}".format(reg)] = listaNAP[0][4]
                ws["I{}".format(reg)] = listaNAP[0][5]
                ws["J{}".format(reg)] = listaNAP[0][6]
                ws["K{}".format(reg)] = listaNAP[0][7]
                ws["L{}".format(reg)] = listaNAP[0][8]
                ws["M{}".format(reg)] = listaNAP[0][9]

            listaTAP = []
            with arcpy.da.SearchCursor(tapfl, [codTAP, "SHAPE@X", "SHAPE@Y", "MTCODNOD", "MTTIPTRO", "MTTRONCAL", "COD_TAP", "MTNUMBOR", "MTCNTBORLBR", "MTTIPO", "NUMCOO_X", "NUMCOO_Y"]) as cursor:
                for n in cursor:
                    cli_fd = [[x[0], x[1], x[2]] for x in arcpy.da.SearchCursor(cli, ["ID", "SHAPE@X", "SHAPE@Y"])]
                    cli_x  = cli_fd[0][1]
                    cli_y  = cli_fd[0][2]
                    tap_x  = n[1]
                    tap_y  = n[2]
                    dist = round(math.sqrt((pow(cli_x - tap_x, 2) + pow(cli_y - tap_y,2)))*111110, 1)
                    rowTAP = [n[3], n[4], n[5], [n[3]+n[4]+n[5]][0], "SI", n[6], n[7], n[8], n[9], dist, n[10], n[11]]
                    listaTAP.append(rowTAP)

            listaTAP.sort(key=lambda x: x[9]) # reverse=True
            if len(listaTAP)>0:
                ws["N{}".format(reg)] = listaTAP[0][0]
                ws["O{}".format(reg)] = listaTAP[0][1]
                ws["P{}".format(reg)] = listaTAP[0][2]
                ws["Q{}".format(reg)] = listaTAP[0][3]
                ws["R{}".format(reg)] = listaTAP[0][4]
                ws["S{}".format(reg)] = listaTAP[0][5]
                ws["T{}".format(reg)] = listaTAP[0][6]
                ws["U{}".format(reg)] = listaTAP[0][7]
                ws["V{}".format(reg)] = listaTAP[0][8]
                ws["W{}".format(reg)] = listaTAP[0][9]
                ws["X{}".format(reg)] = listaTAP[0][10]
                ws["Y{}".format(reg)] = listaTAP[0][11]
            if len(listaTAP)>1:
                ws["Z{}".format(reg)] = listaTAP[1][4]
                ws["AA{}".format(reg)] = listaTAP[1][5]
                ws["AB{}".format(reg)] = listaTAP[1][6]
                ws["AC{}".format(reg)] = listaTAP[1][7]
                ws["AD{}".format(reg)] = listaTAP[1][8]
                ws["AE{}".format(reg)] = listaTAP[1][9]
                ws["AF{}".format(reg)] = listaTAP[1][10]
                ws["AG{}".format(reg)] = listaTAP[1][11]
            if len(listaTAP)>2:
                ws["AH{}".format(reg)] = listaTAP[2][4]
                ws["AI{}".format(reg)] = listaTAP[2][5]
                ws["AJ{}".format(reg)] = listaTAP[2][6]
                ws["AK{}".format(reg)] = listaTAP[2][7]
                ws["AL{}".format(reg)] = listaTAP[2][8]
                ws["AM{}".format(reg)] = listaTAP[2][9]
                ws["AN{}".format(reg)] = listaTAP[2][10]
                ws["AO{}".format(reg)] = listaTAP[2][11]
            reg = reg + 1
            print reg

            arcpy.SelectLayerByAttribute_management(napfl, "CLEAR_SELECTION")  #Limpiar seleccion
            arcpy.SelectLayerByAttribute_management(tapfl, "CLEAR_SELECTION")  #Limpiar seleccion
            arcpy.SelectLayerByAttribute_management(ainfl, "CLEAR_SELECTION") #Limpiar seleccion
        except:
            print traceback.format_exc()

def main():
    wb = load_workbook(pathExcel)
    ws = wb["Hoja1"]
    selectLayer(CLIENTE, ws)
    wb.save(pathExcelout)

if __name__ == '__main__':
    main()
    
print datetime.now() - startTime
print "FINALIZADO"